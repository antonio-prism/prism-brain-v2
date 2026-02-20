"""
PRISM Engine — Fallback rate management.

Loads hardcoded base rates from the existing app's seed files and/or
the Risk Catalog spreadsheet. These serve as the last-resort fallback
when dynamic computation fails.

Fallback chain: computed value → last cached value → hardcoded base rate.
"""

import json
import logging
import os
from pathlib import Path

logger = logging.getLogger(__name__)

# Paths relative to the project root
PROJECT_ROOT = Path(__file__).parent.parent
SEED_DIR = PROJECT_ROOT / "frontend" / "data" / "seeds"
CATALOG_PATH = PROJECT_ROOT / "PRISM_Risk_Catalog.xlsx"
FALLBACK_CACHE_PATH = Path(__file__).parent / "data" / "fallback_rates.json"

_fallback_rates: dict[str, float] = {}


def _load_from_seeds() -> dict[str, float]:
    """Load base rates from the app's seed JSON files."""
    rates = {}
    seed_files = [
        "physical_domain_seed.json",
        "digital_domain_seed.json",
        "structural_domain_seed.json",
        "operational_domain_seed.json",
    ]
    for filename in seed_files:
        path = SEED_DIR / filename
        if not path.exists():
            logger.warning(f"Seed file not found: {path}")
            continue
        try:
            with open(path, "r", encoding="utf-8") as f:
                events = json.load(f)
            for event in events:
                event_id = event.get("event_id")
                # base_rate_pct is stored as percentage (e.g., 0.12 means 0.12%)
                base_pct = event.get("base_rate_pct")
                if event_id and base_pct is not None:
                    rates[event_id] = float(base_pct) / 100.0  # Convert to decimal
        except Exception as e:
            logger.error(f"Error loading seed file {filename}: {e}")
    return rates


def _load_from_catalog() -> dict[str, float]:
    """Load base rates from the Risk Catalog Excel spreadsheet."""
    rates = {}
    if not CATALOG_PATH.exists():
        logger.info("Risk Catalog spreadsheet not found, skipping")
        return rates
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(CATALOG_PATH), data_only=True)
        # Try common sheet names
        sheet_names_to_try = ["Risk Catalog", "Sheet1", "Events"]
        ws = None
        for name in sheet_names_to_try:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            if row[0] and row[5] is not None:
                try:
                    rates[str(row[0])] = float(row[5])
                except (ValueError, TypeError):
                    pass
    except ImportError:
        logger.warning("openpyxl not installed, cannot read Risk Catalog")
    except Exception as e:
        logger.error(f"Error reading Risk Catalog: {e}")
    return rates


def load_fallback_rates() -> dict[str, float]:
    """
    Load all fallback rates. Priority: seed files > Risk Catalog > cached file.
    Returns dict mapping event_id → annual probability (decimal 0-1).
    """
    global _fallback_rates
    if _fallback_rates:
        return _fallback_rates

    # Try seeds first (primary source)
    rates = _load_from_seeds()
    logger.info(f"Loaded {len(rates)} fallback rates from seed files")

    # Fill gaps from Risk Catalog
    if CATALOG_PATH.exists():
        catalog_rates = _load_from_catalog()
        for eid, rate in catalog_rates.items():
            if eid not in rates:
                rates[eid] = rate
        logger.info(f"Loaded {len(catalog_rates)} rates from Risk Catalog ({len(rates)} total)")

    # Save for quick reload
    try:
        with open(FALLBACK_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(rates, f, indent=2)
    except Exception as e:
        logger.warning(f"Could not cache fallback rates: {e}")

    _fallback_rates = rates
    return rates


def get_fallback_rate(event_id: str) -> float:
    """Get the hardcoded fallback rate for a single event."""
    rates = load_fallback_rates()
    rate = rates.get(event_id)
    if rate is None:
        logger.warning(f"No fallback rate for {event_id}, using 0.05 default")
        return 0.05
    return rate
