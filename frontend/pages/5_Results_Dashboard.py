"""
PRISM Brain - Results Dashboard
================================
Visualize and export risk exposure analysis.
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import sys
from pathlib import Path
from datetime import datetime
from utils.theme import inject_prism_theme

APP_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(APP_DIR))

from utils.constants import CURRENCY_SYMBOLS, RISK_DOMAINS
from utils.helpers import (
    format_currency, format_percentage, get_domain_color, get_domain_icon,
    get_risk_level, get_risk_level_color, export_timestamp
)
from modules.database import (
    get_client, get_all_clients, get_client_processes, get_client_risks,
    get_assessments, get_risk_exposure_summary, calculate_risk_exposure
)

st.set_page_config(page_title="Results Dashboard | PRISM Brain", page_icon="💰", layout="wide")


inject_prism_theme()
if 'current_client_id' not in st.session_state:
    st.session_state.current_client_id = None


def client_selector_sidebar():
    """Client selection sidebar."""
    st.sidebar.header("🏢 Current Client")

    clients = get_all_clients()
    if not clients:
        st.sidebar.warning("No clients created")
        return

    client_names = {c['id']: c['name'] for c in clients}

    selected_id = st.sidebar.selectbox(
        "Select Client",
        options=list(client_names.keys()),
        format_func=lambda x: client_names[x],
        index=list(client_names.keys()).index(st.session_state.current_client_id)
              if st.session_state.current_client_id in client_names else 0
    )

    if selected_id != st.session_state.current_client_id:
        st.session_state.current_client_id = selected_id
        st.rerun()


def executive_summary():
    """Executive summary with key metrics."""
    st.subheader("📊 Executive Summary")

    if not st.session_state.current_client_id:
        st.warning("Please select a client")
        return

    client = get_client(st.session_state.current_client_id)
    cid = st.session_state.current_client_id

    # ── Workflow progress check ──
    processes = get_client_processes(cid)
    risks = get_client_risks(cid, prioritized_only=True)
    raw_assessments = get_assessments(cid) or []

    step1_ok = len(processes) > 0
    step2_ok = len(risks) > 0
    step3_ok = len(raw_assessments) > 0
    has_criticality = any(
        (p.get("criticality_per_day") or 0) > 0 for p in processes
    ) if processes else False

    if not (step1_ok and step2_ok and step3_ok):
        st.markdown("### 📋 Data Readiness")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            icon = "✅" if step1_ok else "❌"
            st.markdown(f"{icon} **Processes:** {len(processes)} selected")
        with c2:
            icon = "✅" if has_criticality else "⚠️"
            st.markdown(f"{icon} **Revenue Impact:** {'set' if has_criticality else 'not set'}")
        with c3:
            icon = "✅" if step2_ok else "❌"
            st.markdown(f"{icon} **Risks:** {len(risks)} selected")
        with c4:
            icon = "✅" if step3_ok else "❌"
            st.markdown(f"{icon} **Assessments:** {len(raw_assessments)} completed")

        if not step1_ok:
            st.warning("Step 1: Go to **Process Criticality** and select processes.")
        elif not has_criticality:
            st.warning("Step 2: Go to **Process Criticality → Criticality tab** and set Daily Downtime Revenue Impact values.")
        elif not step2_ok:
            st.warning("Step 3: Go to **Risk Selection** and select risks.")
        elif not step3_ok:
            st.warning("Step 4: Go to **Risk Assessment** and complete at least one process-risk assessment.")
        return

    summary = get_risk_exposure_summary(cid)

    if not summary:
        st.warning("No assessments completed yet. Please complete risk assessments first.")
        return

    currency = client.get('currency', 'EUR')
    symbol = CURRENCY_SYMBOLS.get(currency, '€')

    # Top metrics
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric(
            "Total Annual Risk Exposure",
            format_currency(summary['total_exposure'], currency),
            help="Sum of all assessed risk exposures"
        )

    with col2:
        revenue = client.get('revenue', 0)
        if revenue > 0:
            risk_ratio = summary['total_exposure'] / revenue * 100
            st.metric(
                "% of Revenue at Risk",
                f"{risk_ratio:.1f}%",
                help="Total exposure as percentage of annual revenue"
            )
        else:
            st.metric("Assessments", len(summary['assessments']))

    with col3:
        # Highest risk process
        if summary['by_process']:
            top_process = max(summary['by_process'].items(), key=lambda x: x[1])
            st.metric(
                "Highest Risk Process",
                top_process[0][:20],
                format_currency(top_process[1], currency)
            )

    with col4:
        # Highest risk domain
        if summary['by_domain']:
            top_domain = max(summary['by_domain'].items(), key=lambda x: x[1])
            st.metric(
                "Highest Risk Domain",
                f"{get_domain_icon(top_domain[0])} {top_domain[0]}",
                format_currency(top_domain[1], currency)
            )

    # Client info card
    st.divider()
    st.markdown(f"""
    **Client:** {client['name']} | **Location:** {client.get('location', 'N/A')} |
    **Industry:** {client.get('industry', 'N/A')} | **Currency:** {currency}
    """)


def domain_breakdown_chart(summary, currency):
    """Create domain breakdown pie/bar chart."""
    if not summary['by_domain']:
        return None

    df = pd.DataFrame([
        {"Domain": domain, "Exposure": exposure,
         "Color": get_domain_color(domain)}
        for domain, exposure in summary['by_domain'].items()
    ])

    # Pie chart
    fig = px.pie(
        df, values='Exposure', names='Domain',
        title='Risk Exposure by Domain',
        color='Domain',
        color_discrete_map={d: get_domain_color(d) for d in RISK_DOMAINS.keys()},
        hole=0.4
    )

    fig.update_traces(textposition='inside', textinfo='percent+label')
    fig.update_layout(showlegend=True, height=400)

    return fig


def process_exposure_chart(summary, currency, top_n=10):
    """Create top processes bar chart."""
    if not summary['by_process']:
        return None

    # Sort and take top N
    sorted_procs = sorted(summary['by_process'].items(),
                         key=lambda x: x[1], reverse=True)[:top_n]

    df = pd.DataFrame([
        {"Process": name[:25], "Exposure": exposure}
        for name, exposure in sorted_procs
    ])

    fig = px.bar(
        df, x='Exposure', y='Process',
        orientation='h',
        title=f'Top {top_n} Processes by Risk Exposure',
        color='Exposure',
        color_continuous_scale='Reds'
    )

    fig.update_layout(
        yaxis={'categoryorder': 'total ascending'},
        height=400,
        showlegend=False
    )

    return fig


def risk_exposure_chart(summary, currency, top_n=10):
    """Create top risks bar chart."""
    if not summary['by_risk']:
        return None

    sorted_risks = sorted(summary['by_risk'].items(),
                         key=lambda x: x[1], reverse=True)[:top_n]

    df = pd.DataFrame([
        {"Risk": name[:30], "Exposure": exposure}
        for name, exposure in sorted_risks
    ])

    fig = px.bar(
        df, x='Exposure', y='Risk',
        orientation='h',
        title=f'Top {top_n} Risks by Exposure',
        color='Exposure',
        color_continuous_scale='Oranges'
    )

    fig.update_layout(
        yaxis={'categoryorder': 'total ascending'},
        height=400,
        showlegend=False
    )

    return fig


def heatmap_chart(summary, currency):
    """Create process-domain heatmap."""
    if not summary['assessments']:
        return None

    symbol = CURRENCY_SYMBOLS.get(currency, '€')

    # Aggregate by process and domain
    heatmap_data = {}
    actual_domains = set()
    for a in summary['assessments']:
        process = a['process_name'][:25]
        domain = a['domain']
        exposure = a['exposure']
        actual_domains.add(domain)

        if process not in heatmap_data:
            heatmap_data[process] = {}
        if domain not in heatmap_data[process]:
            heatmap_data[process][domain] = 0
        heatmap_data[process][domain] += exposure

    if not heatmap_data:
        return None

    # Use actual domains from data (in case they don't match RISK_DOMAINS keys)
    domains = sorted(actual_domains)
    processes = list(heatmap_data.keys())

    z_data = []
    for proc in processes:
        row = [heatmap_data.get(proc, {}).get(d, 0) for d in domains]
        z_data.append(row)

    fig = go.Figure(data=go.Heatmap(
        z=z_data,
        x=domains,
        y=processes,
        colorscale='RdYlGn_r',
        text=[[f"{symbol}{v:,.0f}" for v in row] for row in z_data],
        texttemplate="%{text}",
        textfont={"size": 10},
    ))

    fig.update_layout(
        title='Risk Heatmap: Process × Domain',
        height=max(400, len(processes) * 40 + 100),
        xaxis_title='Domain',
        yaxis_title='Process'
    )

    return fig


def visualizations_tab():
    """Visualizations tab content."""
    st.subheader("📈 Risk Visualizations")

    if not st.session_state.current_client_id:
        return

    client = get_client(st.session_state.current_client_id)
    summary = get_risk_exposure_summary(st.session_state.current_client_id)

    if not summary:
        st.warning("No data to visualize. Please complete assessments first.")
        return

    currency = client.get('currency', 'EUR')

    # Charts layout
    col1, col2 = st.columns(2)

    with col1:
        # Domain breakdown
        fig1 = domain_breakdown_chart(summary, currency)
        if fig1:
            st.plotly_chart(fig1, use_container_width=True)

        # Top risks
        fig3 = risk_exposure_chart(summary, currency)
        if fig3:
            st.plotly_chart(fig3, use_container_width=True)

    with col2:
        # Top processes
        fig2 = process_exposure_chart(summary, currency)
        if fig2:
            st.plotly_chart(fig2, use_container_width=True)

        # Domain totals bar
        if summary['by_domain']:
            df_domains = pd.DataFrame([
                {"Domain": d, "Exposure": e, "Icon": get_domain_icon(d)}
                for d, e in summary['by_domain'].items()
            ])

            fig4 = px.bar(
                df_domains, x='Domain', y='Exposure',
                title='Total Exposure by Domain',
                color='Domain',
                color_discrete_map={d: get_domain_color(d) for d in RISK_DOMAINS.keys()}
            )
            fig4.update_layout(showlegend=False, height=350)
            st.plotly_chart(fig4, use_container_width=True)

    # Full heatmap
    st.divider()
    fig_heatmap = heatmap_chart(summary, currency)
    if fig_heatmap:
        st.plotly_chart(fig_heatmap, use_container_width=True)


def detailed_results_tab():
    """Detailed results table."""
    st.subheader("📋 Detailed Results")

    if not st.session_state.current_client_id:
        return

    client = get_client(st.session_state.current_client_id)
    summary = get_risk_exposure_summary(st.session_state.current_client_id)

    if not summary:
        st.warning("No results available.")
        return

    currency = client.get('currency', 'EUR')
    symbol = CURRENCY_SYMBOLS.get(currency, '€')

    # Filter options
    col1, col2, col3 = st.columns(3)
    with col1:
        domain_filter = st.selectbox(
            "Filter by Domain",
            options=["All"] + list(RISK_DOMAINS.keys())
        )
    with col2:
        sort_by = st.selectbox(
            "Sort by",
            options=["Exposure (High to Low)", "Exposure (Low to High)",
                    "Process Name", "Risk Name"]
        )
    with col3:
        min_exposure = st.number_input(
            f"Min Exposure ({symbol})",
            min_value=0,
            value=0,
            step=1000
        )

    # Prepare data
    data = []
    for a in summary['assessments']:
        data.append({
            "Process": a['process_name'],
            "Risk": a['risk_name'],
            "Domain": a['domain'],
            f"Criticality ({symbol}/day)": a['criticality_per_day'],
            "Vulnerability (%)": a['vulnerability'] * 100,
            "Resilience (%)": a['resilience'] * 100,
            "Downtime (days)": a['expected_downtime'],
            "Probability (%)": a['probability'] * 100,
            f"Exposure ({symbol}/yr)": a['exposure']
        })

    df = pd.DataFrame(data)

    # Apply filters
    if domain_filter != "All":
        df = df[df['Domain'] == domain_filter]

    df = df[df[f"Exposure ({symbol}/yr)"] >= min_exposure]

    # Sort
    if sort_by == "Exposure (High to Low)":
        df = df.sort_values(f"Exposure ({symbol}/yr)", ascending=False)
    elif sort_by == "Exposure (Low to High)":
        df = df.sort_values(f"Exposure ({symbol}/yr)", ascending=True)
    elif sort_by == "Process Name":
        df = df.sort_values("Process")
    elif sort_by == "Risk Name":
        df = df.sort_values("Risk")

    # Display
    st.dataframe(
        df,
        use_container_width=True,
        hide_index=True,
        column_config={
            f"Criticality ({symbol}/day)": st.column_config.NumberColumn(format=f"{symbol}%d"),
            "Vulnerability (%)": st.column_config.NumberColumn(format="%.0f%%"),
            "Resilience (%)": st.column_config.NumberColumn(format="%.0f%%"),
            "Probability (%)": st.column_config.NumberColumn(format="%.1f%%"),
            f"Exposure ({symbol}/yr)": st.column_config.NumberColumn(format=f"{symbol}%,.0f"),
        }
    )

    # Summary stats
    st.divider()
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Rows Shown", len(df))
    with col2:
        st.metric("Total Exposure (Filtered)", format_currency(df[f"Exposure ({symbol}/yr)"].sum(), currency))
    with col3:
        st.metric("Avg Exposure per Combination", format_currency(df[f"Exposure ({symbol}/yr)"].mean(), currency))


def export_tab():
    """Export functionality."""
    st.subheader("📥 Export Results")

    if not st.session_state.current_client_id:
        return

    client = get_client(st.session_state.current_client_id)
    summary = get_risk_exposure_summary(st.session_state.current_client_id)

    if not summary:
        st.warning("No data to export.")
        return

    currency = client.get('currency', 'EUR')
    symbol = CURRENCY_SYMBOLS.get(currency, '€')

    st.markdown("""
    Export your risk assessment results in various formats.
    """)

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### 📊 Excel Export")
        st.markdown("Comprehensive workbook with all data and calculations.")

        if st.button("Generate Excel Report", type="primary", use_container_width=True):
            with st.spinner("Generating Excel report..."):
                excel_data = generate_excel_report(client, summary, currency)
                st.download_button(
                    label="📥 Download Excel",
                    data=excel_data,
                    file_name=f"PRISM_Brain_{client['name'].replace(' ', '_')}_{export_timestamp()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    with col2:
        st.markdown("### 📄 CSV Export")
        st.markdown("Raw data for further analysis.")

        if st.button("Generate CSV", use_container_width=True):
            csv_data = generate_csv_report(summary, currency)
            st.download_button(
                label="📥 Download CSV",
                data=csv_data,
                file_name=f"PRISM_Brain_{client['name'].replace(' ', '_')}_{export_timestamp()}.csv",
                mime="text/csv",
                use_container_width=True
            )


def _auto_width(ws, min_width=10, max_width=45):
    """Auto-fit column widths for a worksheet."""
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        # Skip merged cells which lack column_letter attribute
        first_cell = col[0]
        if isinstance(first_cell, MergedCell):
            continue
        col_letter = first_cell.column_letter
        lengths = []
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            try:
                lengths.append(len(str(cell.value or "")))
            except Exception:
                pass
        best = max(lengths) if lengths else min_width
        ws.column_dimensions[col_letter].width = max(min_width, min(best + 3, max_width))


def generate_excel_report(client, summary, currency):
    """Generate a professional 3-tab Excel report.

    Tab 1 — Company Profile: client info from Client Setup.
    Tab 2 — Risk Overview: summary tables by domain, process, risk.
    Tab 3 — Detailed Results: full process × risk matrix.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter

    symbol = CURRENCY_SYMBOLS.get(currency, '€')
    cur_fmt = f'#,##0'
    pct_fmt = '0.0%'

    # ── Shared styles ──
    PRISM_BLUE = "1F4E79"
    PRISM_LIGHT = "D6E4F0"
    PRISM_ACCENT = "2E75B6"

    title_font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor=PRISM_BLUE)
    section_font = Font(name="Arial", bold=True, size=12, color=PRISM_BLUE)
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=PRISM_ACCENT)
    label_font = Font(name="Arial", bold=True, size=10)
    value_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", fgColor=PRISM_LIGHT)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC")
    )

    def _write_title_row(ws, text, row=1, cols=6):
        """Merge cells and write a coloured title bar."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 32
        for c in range(2, cols + 1):
            ws.cell(row=row, column=c).fill = title_fill

    def _write_header_row(ws, headers, row):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[row].height = 22

    wb = Workbook()

    # ================================================================
    # TAB 1 — COMPANY PROFILE
    # ================================================================
    ws1 = wb.active
    ws1.title = "Company Profile"
    ws1.sheet_properties.tabColor = PRISM_BLUE

    _write_title_row(ws1, f"PRISM Brain — Risk Report: {client['name']}", row=1, cols=4)

    # Report metadata
    ws1.cell(row=3, column=1, value="Report Generated").font = label_font
    ws1.cell(row=3, column=2, value=datetime.now().strftime("%d %B %Y, %H:%M")).font = value_font

    # Company details — label/value pairs
    fields = [
        ("Company Name", client.get("name", "")),
        ("Location", client.get("location", "")),
        ("Industry", client.get("industry", "")),
        ("Sectors", client.get("sectors", "")),
        ("Annual Revenue", client.get("revenue", 0)),
        ("Employees", client.get("employees", 0)),
        ("Currency", currency),
        ("Export Percentage", client.get("export_percentage", 0)),
        ("Primary Markets", client.get("primary_markets", "")),
        ("Notes", client.get("notes", "")),
    ]

    ws1.cell(row=5, column=1, value="Company Information").font = section_font
    r = 6
    for label, val in fields:
        ws1.cell(row=r, column=1, value=label).font = label_font
        c = ws1.cell(row=r, column=2, value=val)
        c.font = value_font
        if label == "Annual Revenue":
            c.number_format = f'{symbol}#,##0'
        elif label == "Export Percentage":
            c.number_format = '0%'
        ws1.cell(row=r, column=1).border = thin_border
        ws1.cell(row=r, column=2).border = thin_border
        r += 1

    # Key metrics section
    r += 1
    ws1.cell(row=r, column=1, value="Risk Exposure Summary").font = section_font
    r += 1

    metrics = [
        ("Total Annual Risk Exposure", summary["total_exposure"], f'{symbol}#,##0'),
        ("Processes Assessed", len(summary.get("by_process", {})), '#,##0'),
        ("Risks Assessed", len(summary.get("by_risk", {})), '#,##0'),
        ("Domains Covered", len(summary.get("by_domain", {})), '#,##0'),
        ("Total Assessments", len(summary.get("assessments", [])), '#,##0'),
    ]

    revenue = client.get("revenue", 0)
    if revenue and revenue > 0:
        ratio = summary["total_exposure"] / revenue
        metrics.insert(1, ("% of Revenue at Risk", ratio, '0.0%'))

    for label, val, fmt in metrics:
        ws1.cell(row=r, column=1, value=label).font = label_font
        c = ws1.cell(row=r, column=2, value=val)
        c.font = Font(name="Arial", bold=True, size=11, color=PRISM_BLUE)
        c.number_format = fmt
        ws1.cell(row=r, column=1).border = thin_border
        ws1.cell(row=r, column=2).border = thin_border
        r += 1

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 35

    # ================================================================
    # TAB 2 — RISK OVERVIEW (data for visualisation)
    # ================================================================
    ws2 = wb.create_sheet("Risk Overview")
    ws2.sheet_properties.tabColor = "2E75B6"

    _write_title_row(ws2, "Risk Exposure Overview", row=1, cols=6)

    # Section A: Exposure by Domain
    ws2.cell(row=3, column=1, value="Exposure by Domain").font = section_font
    _write_header_row(ws2, ["Domain", f"Exposure ({symbol}/yr)", "% of Total"], row=4)

    total_exp = summary["total_exposure"] or 1
    r = 5
    domain_start = r
    for domain, exposure in sorted(summary["by_domain"].items(),
                                    key=lambda x: x[1], reverse=True):
        ws2.cell(row=r, column=1, value=domain).font = value_font
        ws2.cell(row=r, column=2, value=exposure).number_format = cur_fmt
        ws2.cell(row=r, column=3, value=exposure / total_exp).number_format = pct_fmt
        if (r - domain_start) % 2 == 1:
            for c in range(1, 4):
                ws2.cell(row=r, column=c).fill = alt_fill
        r += 1
    domain_end = r - 1

    # Pie chart for domain breakdown
    if domain_end >= domain_start:
        pie = PieChart()
        pie.title = "Exposure by Domain"
        pie.style = 10
        data_ref = Reference(ws2, min_col=2, min_row=4, max_row=domain_end)
        cats_ref = Reference(ws2, min_col=1, min_row=5, max_row=domain_end)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        pie.width = 16
        pie.height = 10
        ws2.add_chart(pie, f"E3")

    # Section B: Top Processes
    r += 2
    ws2.cell(row=r, column=1, value="Top Processes by Exposure").font = section_font
    r += 1
    _write_header_row(ws2, ["Process", f"Exposure ({symbol}/yr)", "% of Total"], row=r)
    r += 1
    proc_start = r

    sorted_procs = sorted(summary["by_process"].items(),
                           key=lambda x: x[1], reverse=True)[:15]
    for proc_name, exposure in sorted_procs:
        ws2.cell(row=r, column=1, value=proc_name).font = value_font
        ws2.cell(row=r, column=2, value=exposure).number_format = cur_fmt
        ws2.cell(row=r, column=3, value=exposure / total_exp).number_format = pct_fmt
        if (r - proc_start) % 2 == 1:
            for c in range(1, 4):
                ws2.cell(row=r, column=c).fill = alt_fill
        r += 1
    proc_end = r - 1

    # Bar chart for top processes
    if proc_end >= proc_start:
        bar = BarChart()
        bar.type = "bar"
        bar.title = "Top Processes by Exposure"
        bar.style = 10
        bar.y_axis.title = None
        bar.x_axis.title = f"Exposure ({symbol}/yr)"
        data_ref = Reference(ws2, min_col=2, min_row=proc_start - 1, max_row=proc_end)
        cats_ref = Reference(ws2, min_col=1, min_row=proc_start, max_row=proc_end)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.width = 18
        bar.height = 12
        ws2.add_chart(bar, f"E{proc_start - 1}")

    # Section C: Top Risks
    r += 2
    ws2.cell(row=r, column=1, value="Top Risks by Exposure").font = section_font
    r += 1
    _write_header_row(ws2, ["Risk", f"Exposure ({symbol}/yr)", "% of Total"], row=r)
    r += 1
    risk_start = r

    sorted_risks = sorted(summary["by_risk"].items(),
                            key=lambda x: x[1], reverse=True)[:15]
    for risk_name, exposure in sorted_risks:
        ws2.cell(row=r, column=1, value=risk_name).font = value_font
        ws2.cell(row=r, column=2, value=exposure).number_format = cur_fmt
        ws2.cell(row=r, column=3, value=exposure / total_exp).number_format = pct_fmt
        if (r - risk_start) % 2 == 1:
            for c in range(1, 4):
                ws2.cell(row=r, column=c).fill = alt_fill
        r += 1

    _auto_width(ws2)

    # ================================================================
    # TAB 3 — DETAILED RESULTS
    # ================================================================
    ws3 = wb.create_sheet("Detailed Results")
    ws3.sheet_properties.tabColor = "70AD47"

    _write_title_row(ws3, "Detailed Risk Assessment Results", row=1, cols=9)

    detail_headers = [
        "Process", "Risk", "Domain",
        f"Criticality ({symbol}/day)", "Vulnerability",
        "Resilience", "Downtime (days)",
        "Probability", f"Exposure ({symbol}/yr)"
    ]
    _write_header_row(ws3, detail_headers, row=3)

    # Freeze top rows so headers stay visible when scrolling
    ws3.freeze_panes = "A4"

    for i, a in enumerate(summary["assessments"]):
        r = i + 4
        ws3.cell(row=r, column=1, value=a["process_name"]).font = value_font
        ws3.cell(row=r, column=2, value=a["risk_name"]).font = value_font
        ws3.cell(row=r, column=3, value=a["domain"]).font = value_font
        ws3.cell(row=r, column=4, value=a["criticality_per_day"]).number_format = cur_fmt
        ws3.cell(row=r, column=5, value=a["vulnerability"]).number_format = pct_fmt
        ws3.cell(row=r, column=6, value=a["resilience"]).number_format = pct_fmt
        ws3.cell(row=r, column=7, value=a["expected_downtime"])
        ws3.cell(row=r, column=8, value=a["probability"]).number_format = pct_fmt
        ws3.cell(row=r, column=9, value=a["exposure"]).number_format = cur_fmt
        # Alternating row shading
        if i % 2 == 1:
            for c in range(1, 10):
                ws3.cell(row=r, column=c).fill = alt_fill

    # Totals row
    n = len(summary["assessments"])
    if n > 0:
        total_row = n + 4
        ws3.cell(row=total_row, column=1, value="TOTAL").font = Font(
            name="Arial", bold=True, size=10)
        total_cell = ws3.cell(row=total_row, column=9,
                               value=f"=SUM(I4:I{total_row - 1})")
        total_cell.number_format = cur_fmt
        total_cell.font = Font(name="Arial", bold=True, size=10, color=PRISM_BLUE)

    _auto_width(ws3)

    # ── Save ──
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_csv_report(summary, currency):
    """Generate CSV report."""
    symbol = CURRENCY_SYMBOLS.get(currency, '€')

    data = []
    for a in summary['assessments']:
        data.append({
            "Process": a['process_name'],
            "Risk": a['risk_name'],
            "Domain": a['domain'],
            "Criticality_per_day": a['criticality_per_day'],
            "Vulnerability": a['vulnerability'],
            "Resilience": a['resilience'],
            "Downtime_days": a['expected_downtime'],
            "Probability": a['probability'],
            "Exposure_annual": a['exposure']
        })

    df = pd.DataFrame(data)
    return df.to_csv(index=False).encode('utf-8')


def main():
    """Main page function."""
    st.title("💰 Results Dashboard")
    st.markdown("View and export your risk exposure analysis.")

    client_selector_sidebar()

    if not st.session_state.current_client_id:
        st.warning("Please select a client to view results")
        if st.button("Go to Client Setup"):
            st.switch_page("pages/1_Client_Setup.py")
        return

    # Executive summary at top
    executive_summary()

    st.divider()

    # Tabs for different views
    tab1, tab2, tab3 = st.tabs([
        "📈 Visualizations",
        "📋 Detailed Results",
        "📥 Export"
    ])

    with tab1:
        visualizations_tab()

    with tab2:
        detailed_results_tab()

    with tab3:
        export_tab()

    # Navigation
    st.divider()
    col1, col2, col3 = st.columns([1, 2, 1])

    with col1:
        if st.button("← Risk Assessment"):
            st.switch_page("pages/4_Risk_Assessment.py")

    with col3:
        if st.button("🏠 Home"):
            st.switch_page("Welcome.py")


if __name__ == "__main__":
    main()
