"""
PRISM Brain V2 - Risk Selection Module
=======================================
Select and prioritize risks relevant to the client.
Uses V2 taxonomy: Domains -> Families -> Events with base-rate probabilities.
Events are grouped by Domain and Family with collapsible sections.
"""

import streamlit as st
import pandas as pd
import sys
import io
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.theme import inject_prism_theme

APP_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(APP_DIR))

from utils.constants import RISK_DOMAINS
from utils.helpers import (
    get_domain_icon,
    format_percentage
)
from modules.database import (
    get_client,
    get_all_clients,
    add_client_risk,
    get_client_risks,
    is_backend_online
)
from modules.api_client import (
    fetch_v2_events_normalized,
    api_v2_health,
    api_engine_compute_all,
)

st.set_page_config(
    page_title="Risk Selection | PRISM Brain",
    page_icon="\u26A1",
    layout="wide"
)

inject_prism_theme()


# ====================== Session State ======================
if 'current_client_id' not in st.session_state:
    st.session_state.current_client_id = None

if 'selected_risks' not in st.session_state:
    st.session_state.selected_risks = set()

if 'engine_results' not in st.session_state:
    st.session_state.engine_results = None

if 'active_family' not in st.session_state:
    st.session_state.active_family = None

if '_risks_synced_for' not in st.session_state:
    st.session_state._risks_synced_for = None

# Domain display order (matches family code numbering: 1.x, 2.x, 3.x, 4.x)
_DOMAIN_ORDER = [
    ("PHYSICAL", "Physical"),
    ("STRUCTURAL", "Structural"),
    ("DIGITAL", "Digital"),
    ("OPERATIONAL", "Operational"),
]


# ====================== Engine Results Helper ======================

def _get_engine_results_if_cached():
    """Return engine results from session state if already computed.

    NEVER triggers a new engine computation — returns None if no results
    are available.  This keeps the page responsive.
    """
    return st.session_state.get('engine_results')


def _get_probability_for_risk(risk, engine_results=None):
    """Get the best probability for a risk.

    Uses engine result if available, otherwise falls back to base rate.
    Returns (probability_0_to_1, source_label).
    """
    if engine_results:
        engine_data = engine_results.get(risk['id'])
        if engine_data and isinstance(engine_data, dict):
            prob = engine_data.get("layer1", {}).get("p_global", 0)
            return prob, "Engine"
    return risk.get('default_probability', 0), "Base Rate"


def _set_active_family(family_code):
    """Callback: remember which family expander the user is working in."""
    st.session_state.active_family = family_code


def _count_selected_in_risks(risk_list):
    """Count how many risks from a list are currently selected."""
    count = 0
    for r in risk_list:
        wkey = f"risk_{r['id']}"
        if wkey in st.session_state:
            if st.session_state[wkey]:
                count += 1
        elif r['id'] in st.session_state.selected_risks:
            count += 1
    return count


# ====================== Sidebar ======================

def client_selector_sidebar():
    """Sidebar for client selection and V2 backend status."""
    st.sidebar.header("\U0001F3E2 Current Client")
    clients = get_all_clients()

    if not clients:
        st.sidebar.warning("No clients created yet")
        if st.sidebar.button("Create Client"):
            st.switch_page("pages/1_Client_Setup.py")
        return

    client_names = {c['id']: c['name'] for c in clients}
    selected_id = st.sidebar.selectbox(
        "Select Client",
        options=list(client_names.keys()),
        format_func=lambda x: client_names[x],
        index=list(client_names.keys()).index(st.session_state.current_client_id)
        if st.session_state.current_client_id in client_names
        else 0
    )

    if selected_id != st.session_state.current_client_id:
        st.session_state.current_client_id = selected_id
        existing_risks = get_client_risks(selected_id, prioritized_only=True)
        st.session_state.selected_risks = set(r['risk_id'] for r in existing_risks)
        st.rerun()

    if st.session_state.current_client_id:
        client = get_client(st.session_state.current_client_id)
        st.sidebar.divider()
        st.sidebar.markdown(f"**\U0001F4CD {client.get('location', 'N/A')}**")
        st.sidebar.markdown(f"\U0001F3ED {client.get('industry', 'N/A')}")
        st.sidebar.markdown(f"\U0001F4CA {client.get('sectors', 'N/A')}")

    # V2 backend status
    st.sidebar.divider()
    st.sidebar.caption("**V2 Backend**")
    if is_backend_online():
        v2h = api_v2_health()
        if v2h and v2h.get("status") == "healthy":
            st.sidebar.markdown(f"\U0001F7E2 **Connected** \u2014 {v2h.get('v2_events', 0)} events")
        else:
            st.sidebar.markdown("\U0001F7E1 **Connected** \u2014 No V2 events loaded")
    else:
        st.sidebar.markdown("\U0001F534 **Backend offline**")


# ====================== Load V2 Events ======================

def load_risks():
    """Load V2 events from the backend API, normalized for frontend use.

    Returns list of dicts with keys: id, name, domain, family_code,
    family_name, default_probability (0-1), base_rate_pct (0-100), etc.
    """
    return fetch_v2_events_normalized()


# ====================== Risk Selection Tab ======================

def risk_selection_interface():
    """Risk selection interface grouped by Domain and Family.

    Structure: Domain header -> Family expanders -> Event checkboxes.
    Same pattern as Process Criticality (Scope -> Macro-process -> Sub-process).
    """
    client = get_client(st.session_state.current_client_id)
    risks = load_risks()

    if not risks:
        st.error("Could not load V2 events from the backend. "
                 "Make sure the backend is running.")
        return

    # Sync risk_* widget keys from selected_risks BEFORE any checkbox renders.
    # Runs once after import, client change, or bulk action (when flag is reset).
    if st.session_state._risks_synced_for != st.session_state.current_client_id:
        for risk in risks:
            st.session_state[f"risk_{risk['id']}"] = risk['id'] in st.session_state.selected_risks
        st.session_state._risks_synced_for = st.session_state.current_client_id

    st.markdown(f"## Select Risks for {client['name']}")
    st.markdown("Browse risks by domain and family, then check the ones relevant to this client.")

    # ---------- Search + Bulk actions ----------
    search_term = st.text_input("\U0001F50D Search risks", key="risk_search")

    # Filter risks by search term
    if search_term:
        term = search_term.lower()
        filtered_risks = [
            r for r in risks
            if term in r['name'].lower()
            or term in r.get('description', '').lower()
            or term in r.get('id', '').lower()
        ]
    else:
        filtered_risks = risks

    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("\u2713 Select All Visible"):
            for r in filtered_risks:
                st.session_state.selected_risks.add(r['id'])
            st.session_state._risks_synced_for = None
            st.rerun()
    with col2:
        if st.button("\u2717 Clear Selection"):
            st.session_state.selected_risks.clear()
            st.session_state._risks_synced_for = None
            st.rerun()
    with col3:
        total_selected = _count_selected_in_risks(risks)
        st.write(f"**{total_selected}** risks selected")

    st.divider()

    # ---------- Engine results for probability display ----------
    engine_results = _get_engine_results_if_cached()

    # ---------- Group risks by domain and family ----------
    for domain_key, domain_display in _DOMAIN_ORDER:
        domain_info = RISK_DOMAINS.get(domain_key, {})
        domain_icon = domain_info.get("icon", "")
        domain_desc = domain_info.get("description", "")

        # Get filtered risks in this domain
        domain_risks = [
            r for r in filtered_risks
            if r.get('domain', '').upper() == domain_key
        ]

        if not domain_risks:
            continue

        # Group by family
        families = {}
        for r in domain_risks:
            fc = r.get('family_code', '0.0')
            if fc not in families:
                families[fc] = {
                    'name': r.get('family_name', ''),
                    'events': []
                }
            families[fc]['events'].append(r)

        # Count selected in this domain
        domain_selected = _count_selected_in_risks(domain_risks)

        st.markdown(
            f"#### {domain_icon} {domain_key} \u2014 {domain_desc} "
            f"({domain_selected} selected)"
        )

        # Display families sorted by code
        for fc in sorted(families.keys()):
            fam = families[fc]
            events = sorted(fam['events'], key=lambda r: r['id'])

            # Count selected in this family
            fam_selected = _count_selected_in_risks(events)

            # Keep expander open if user just interacted with it
            keep_open = st.session_state.active_family == fc

            with st.expander(
                f"**{fc} \u2014 {fam['name']}** ({fam_selected}/{len(events)} selected)",
                expanded=keep_open,
            ):
                # Select all / Deselect all for this family
                col_a, col_b = st.columns(2)
                with col_a:
                    if st.button("Select all", key=f"selall_{fc}"):
                        st.session_state.active_family = fc
                        for r in events:
                            st.session_state.selected_risks.add(r['id'])
                        st.session_state._risks_synced_for = None
                        st.rerun()
                with col_b:
                    if st.button("Deselect all", key=f"deselall_{fc}"):
                        st.session_state.active_family = fc
                        for r in events:
                            st.session_state.selected_risks.discard(r['id'])
                        st.session_state._risks_synced_for = None
                        st.rerun()

                # Event checkboxes with probability
                for risk in events:
                    risk_id = risk['id']
                    is_selected = risk_id in st.session_state.selected_risks

                    # Probability info
                    prob, source = _get_probability_for_risk(risk, engine_results)
                    prob_str = format_percentage(prob)
                    if source == "Engine":
                        prob_str += " \u25CF"

                    col1, col2 = st.columns([5, 1])
                    with col1:
                        checkbox_kwargs = {
                            "label": f"{risk_id} \u2014 {risk['name']}",
                            "key": f"risk_{risk_id}",
                            "on_change": _set_active_family,
                            "args": (fc,),
                        }
                        if f"risk_{risk_id}" not in st.session_state:
                            checkbox_kwargs["value"] = is_selected
                        if st.checkbox(**checkbox_kwargs):
                            st.session_state.selected_risks.add(risk_id)
                        else:
                            st.session_state.selected_risks.discard(risk_id)
                    with col2:
                        st.write(prob_str)

    # ---------- Save inline ----------
    st.divider()
    if st.button("\U0001F4BE Save Risk Selection", key="save_risks", type="primary"):
        _save_selected_risks(risks)


# ====================== Probabilities Tab ======================

def probability_overview_interface():
    """Show engine-computed probabilities for selected risks."""
    st.subheader("\U0001F4CA Risk Probabilities")

    risks = load_risks()
    selected_risks = [r for r in risks if r['id'] in st.session_state.selected_risks]

    if not selected_risks:
        st.info("Select risks in the 'Select Risks' tab first.")
        return

    engine_results = _get_engine_results_if_cached()

    # Show compute button if engine hasn't been run yet
    if not engine_results:
        st.info("Engine probabilities have not been computed yet. "
                "Click below to compute live probabilities from external data sources. "
                "This takes 1-2 minutes on first run.")
        if st.button("Compute Engine Probabilities", type="primary"):
            with st.spinner("Computing probabilities for 174 events from live data sources... This may take 1-2 minutes."):
                engine_results = api_engine_compute_all(use_cache=False)
                if engine_results:
                    st.session_state.engine_results = engine_results
                    st.rerun()
                else:
                    st.error("Engine computation failed. Check that the backend is running.")
                    return

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Selected Risks", len(selected_risks))
    with col2:
        engine_count = sum(1 for r in selected_risks
                          if engine_results and r['id'] in engine_results)
        st.metric("Engine-Computed", engine_count)
    with col3:
        backend_status = "\U0001F7E2 Online" if is_backend_online() else "\U0001F534 Offline"
        st.write(f"**Backend:** {backend_status}")

    # Refresh button (when engine results already exist)
    if engine_results:
        if st.button("Refresh Probabilities"):
            with st.spinner("Recomputing..."):
                fresh = api_engine_compute_all(use_cache=False)
                if fresh:
                    st.session_state.engine_results = fresh
                    st.rerun()

    st.divider()

    # Build results table with engine data
    results_data = []
    for risk in selected_risks:
        prob, source = _get_probability_for_risk(risk, engine_results)

        if source == "Engine" and engine_results:
            engine_data = engine_results.get(risk['id'], {})
            layer1 = engine_data.get("layer1", {})
            derivation = layer1.get("derivation", {})
            method = layer1.get("method", "?")
            confidence = derivation.get("confidence", "")
        else:
            method = "-"
            confidence = ""

        results_data.append({
            'Event ID': risk['id'],
            'Risk Name': risk['name'],
            'Probability %': round(prob * 100, 2),
            'Method': method,
            'Confidence': confidence,
            'Source': source,
        })

    results_df = pd.DataFrame(results_data)
    results_df = results_df.sort_values('Probability %', ascending=False)

    st.dataframe(
        results_df,
        width="stretch",
        hide_index=True,
        column_config={
            "Probability %": st.column_config.ProgressColumn(
                "Probability %",
                min_value=0,
                max_value=100,
                format="%.1f%%",
            ),
        },
    )

    # Legend
    st.caption(
        "**Method:** A = Frequency count, B = Incidence rate, C = Structural calibration. "
        "**Source:** Engine = computed from live data, Base Rate = raw seed value."
    )


# ====================== Save Tab ======================

def save_risks_interface():
    """Review and save selected risks to the client's portfolio."""
    st.subheader("\U0001F4BE Save Risk Selections")

    if not st.session_state.selected_risks:
        st.warning("No risks selected yet. Use the 'Select Risks' tab first.")
        return

    client = get_client(st.session_state.current_client_id)
    risks = load_risks()
    selected_risks = [r for r in risks if r['id'] in st.session_state.selected_risks]

    st.write(f"**Client:** {client['name']}  \u2014  **Selected Risks:** {len(selected_risks)}")

    # Summary table — use cached engine results if available
    engine_results = _get_engine_results_if_cached()
    selected_df = pd.DataFrame([
        {
            'Event ID': r['id'],
            'Risk Name': r['name'],
            'Domain': r['domain'],
            'Family': r.get('family_name', ''),
            'Probability (%)': f"{_get_probability_for_risk(r, engine_results)[0] * 100:.1f}%"
        }
        for r in selected_risks
    ])
    st.dataframe(selected_df, width="stretch", hide_index=True)

    col1, col2 = st.columns(2)
    with col1:
        if st.button("\u2713 Confirm & Save Risks", type="primary"):
            _save_selected_risks(risks)

    with col2:
        if st.button("\u2190 Back to Risk Selection"):
            st.rerun()


def _save_selected_risks(all_risks):
    """Save all selected risks to the client's risk portfolio.

    Uses engine-computed probability if available in session state,
    otherwise falls back to base_rate_pct.  Probability stored as 0-1
    (exposure formula expects this scale).
    """
    engine_results = _get_engine_results_if_cached()
    saved = 0
    engine_count = 0
    for risk_id in st.session_state.selected_risks:
        risk = next((r for r in all_risks if r['id'] == risk_id), None)
        if risk:
            prob, source = _get_probability_for_risk(risk, engine_results)
            if source == "Engine":
                engine_count += 1

            add_client_risk(
                client_id=st.session_state.current_client_id,
                risk_id=risk_id,
                risk_name=risk['name'],
                domain=risk['domain'],
                category=risk.get('family_name', ''),
                probability=prob,
                is_prioritized=1,
                notes=risk.get('description', ''),
            )
            saved += 1

    msg = f"Saved {saved} risks for this client!"
    if engine_count > 0:
        msg += f" ({engine_count} with engine-computed probabilities)"
    st.success(msg)


# ====================== Import/Export Tab ======================

def import_export_risks():
    """Import and export risk selections as Excel files."""
    st.subheader("\U0001F4E5 Import / Export Risk Selections")

    if not st.session_state.current_client_id:
        st.warning("Select a client first")
        return

    client = get_client(st.session_state.current_client_id)
    risks = load_risks()

    if not risks:
        st.error("Could not load risk events from the backend.")
        return

    col_dl, col_ul = st.columns(2)

    # --- Download template with ALL 174 risks ---
    with col_dl:
        st.markdown("#### Download Risk Template")
        st.markdown(
            "Download a template with **all 174 risk events**. "
            "Review each risk and set the **Selected** column to "
            "**Yes** or **No** to indicate relevance for this client. "
            "Currently selected risks are pre-marked as Yes."
        )

        engine_results = _get_engine_results_if_cached()
        export_data = []
        for risk in sorted(risks, key=lambda r: (r.get('domain', ''), r.get('family_code', ''), r['id'])):
            prob, source = _get_probability_for_risk(risk, engine_results)
            export_data.append({
                'Domain': risk['domain'],
                'Family': risk.get('family_name', ''),
                'Risk ID': risk['id'],
                'Risk Name': risk['name'],
                'Global Probability (%)': round(prob * 100, 2),
                'Selected': 'Yes' if risk['id'] in st.session_state.selected_risks else 'No',
            })

        export_df = pd.DataFrame(export_data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            export_df.to_excel(writer, sheet_name='Risk Selection', index=False)
            ws = writer.sheets['Risk Selection']

            # --- Style definitions ---
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9'),
            )
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill_info = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_fill_input = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
            yes_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            no_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            zebra_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

            num_cols = len(export_df.columns)
            num_rows = len(export_df)
            # "Selected" is the last column
            selected_col_idx = num_cols

            # --- Format header row ---
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                if col_idx == selected_col_idx:
                    cell.fill = header_fill_input
                else:
                    cell.fill = header_fill_info

            # --- Format data rows ---
            for row_idx in range(2, num_rows + 2):
                is_odd = (row_idx % 2) == 1
                selected_val = str(ws.cell(row=row_idx, column=selected_col_idx).value).strip().lower()
                for col_idx in range(1, num_cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')

                    if col_idx == selected_col_idx:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(bold=True)
                        if selected_val == 'yes':
                            cell.fill = yes_fill
                        else:
                            cell.fill = no_fill
                    elif selected_val == 'yes':
                        cell.fill = yes_fill
                    elif is_odd:
                        cell.fill = zebra_fill

            # --- Auto-fit column widths ---
            for col_idx, col_name in enumerate(export_df.columns, 1):
                max_len = max(
                    len(str(col_name)),
                    export_df.iloc[:, col_idx - 1].astype(str).str.len().max()
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

            # --- Freeze header row ---
            ws.freeze_panes = 'A2'
            ws.row_dimensions[1].height = 30

        output.seek(0)

        st.download_button(
            label="Download Risk Template (.xlsx)",
            data=output.getvalue(),
            file_name=f"PRISM_Risk_Selection_{client['name'].replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # --- Upload filled template ---
    with col_ul:
        st.markdown("#### Upload Risk Selection")
        st.markdown(
            "Upload the completed template. Risks with **Selected = Yes** "
            "will replace the current selection."
        )

        uploaded_file = st.file_uploader(
            "Upload completed risk template",
            type=['xlsx'],
            key="risk_upload",
        )

    if uploaded_file:
        try:
            upload_df = pd.read_excel(uploaded_file, sheet_name='Risk Selection')

            valid_ids = {r['id'] for r in risks}
            selected_from_upload = set()

            # Support both old "Event ID" and new "Risk ID" column names
            id_col = 'Risk ID' if 'Risk ID' in upload_df.columns else 'Event ID'
            sel_col = 'Selected' if 'Selected' in upload_df.columns else None

            if id_col not in upload_df.columns or sel_col is None:
                st.error(
                    "File must have columns 'Risk ID' (or 'Event ID') and 'Selected'. "
                    "Please use the PRISM Risk Template."
                )
                return

            for _, row in upload_df.iterrows():
                if str(row.get(sel_col, '')).strip().lower() == 'yes':
                    event_id = str(row.get(id_col, '')).strip()
                    if event_id in valid_ids:
                        selected_from_upload.add(event_id)

            st.info(f"Found **{len(selected_from_upload)}** risks marked as Selected in the file.")

            if selected_from_upload:
                # Preview
                preview_risks = [r for r in risks if r['id'] in selected_from_upload]
                preview_df = pd.DataFrame([
                    {"Domain": r['domain'], "Risk ID": r['id'], "Risk Name": r['name']}
                    for r in sorted(preview_risks, key=lambda r: r['id'])
                ])
                st.dataframe(preview_df, hide_index=True, use_container_width=True)

                if st.button("Apply Import", type="primary", key="apply_risk_upload"):
                    st.session_state.selected_risks = selected_from_upload
                    # Force risk_* widget keys to re-sync from selected_risks
                    # on the next rerun (sync runs BEFORE checkboxes render).
                    st.session_state._risks_synced_for = None
                    st.success(f"Imported {len(selected_from_upload)} risks!")
                    st.rerun()
            else:
                st.warning("No risks marked as 'Yes' in the Selected column.")

        except Exception as e:
            st.error(f"Error reading file: {str(e)}")


# ====================== Main ======================

def main():
    """Main page layout."""
    client_selector_sidebar()

    if not st.session_state.current_client_id:
        st.warning("\U0001F448 Select a client from the sidebar")
        return

    # Header with navigation
    st.title("\u26A1 Risk Selection")

    col1, col2, col3 = st.columns([1, 2, 1])
    with col1:
        if st.button("\u2190 Process Criticality"):
            st.switch_page("pages/2_Process_Criticality.py")
    with col3:
        if st.button("Next: Assessment \u2192"):
            st.switch_page("pages/4_Risk_Assessment.py")

    st.divider()

    # Tabs — Select first (most common action), then Probabilities, Save, Import/Export
    tab1, tab2, tab3, tab4 = st.tabs([
        "\U0001F3AF Select Risks",
        "\U0001F4CA Probabilities",
        "\U0001F4BE Save",
        "\U0001F4E5 Import / Export"
    ])

    with tab1:
        risk_selection_interface()

    with tab2:
        probability_overview_interface()

    with tab3:
        save_risks_interface()

    with tab4:
        import_export_risks()


if __name__ == "__main__":
    main()
