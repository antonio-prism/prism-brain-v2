""" PRISM Brain - Risk Assessment Module
=====================================
Input vulnerability, resilience, and downtime for each process-risk combination.
"""

import streamlit as st
import pandas as pd
import io
import sys
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.theme import inject_prism_theme

APP_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(APP_DIR))

from utils.constants import CURRENCY_SYMBOLS
from utils.helpers import (
    get_domain_color,
    get_domain_icon,
    format_currency,
    format_percentage,
    get_risk_level,
    get_risk_level_color
)
from modules.database import (
    get_client,
    get_all_clients,
    get_client_processes,
    get_client_risks,
    get_assessments,
    save_assessment,
    calculate_risk_exposure
)

st.set_page_config(
    page_title="Risk Assessment | PRISM Brain",
    page_icon="🎯",
    layout="wide"
)

inject_prism_theme()

if 'current_client_id' not in st.session_state:
    st.session_state.current_client_id = None

if 'assessment_mode' not in st.session_state:
    st.session_state.assessment_mode = 'guided'  # 'guided' or 'table'


# ---- Performance: load data ONCE and reuse across all functions ----
def _load_page_data():
    """Load all data needed by this page once, cache in session_state."""
    cid = st.session_state.current_client_id
    if not cid:
        return None, [], [], []
    client = get_client(cid)
    processes = get_client_processes(cid)
    risks = get_client_risks(cid, prioritized_only=True)
    assessments = get_assessments(cid) or []
    return client, processes, risks, assessments


def client_selector_sidebar(processes, risks, assessments):
    """Sidebar for client selection and progress."""
    st.sidebar.header("🏢 Current Client")
    clients = get_all_clients()

    if not clients:
        st.sidebar.warning("No clients created")
        return

    client_names = {c['id']: c['name'] for c in clients}
    selected_id = st.sidebar.selectbox(
        "Select Client",
        options=list(client_names.keys()),
        format_func=lambda x: client_names[x],
        index=list(client_names.keys()).index(st.session_state.current_client_id)
        if st.session_state.current_client_id in client_names
        else 0
    )

    if selected_id != st.session_state.current_client_id:
        st.session_state.current_client_id = selected_id
        st.rerun()

    # Progress indicator (uses data passed in, not re-fetched)
    if st.session_state.current_client_id:
        total_combinations = len(processes) * len(risks)
        completed = len(assessments)

        st.sidebar.divider()
        st.sidebar.subheader("📊 Progress")

        if total_combinations > 0:
            progress = completed / total_combinations
            st.sidebar.progress(progress)
            st.sidebar.write(f"{completed} / {total_combinations} assessed")
            st.sidebar.write(f"({progress*100:.0f}% complete)")
        else:
            st.sidebar.warning("No combinations to assess")


def assessment_overview(client, processes, risks, assessments):
    """Overview of assessment status."""
    st.subheader("📊 Assessment Overview")

    if not st.session_state.current_client_id:
        st.warning("Please select a client")
        return

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Processes", len(processes))
    with col2:
        st.metric("Risks", len(risks))
    with col3:
        total = len(processes) * len(risks)
        st.metric("Combinations", total)
    with col4:
        st.metric("Assessed", len(assessments))

    if not processes:
        st.warning("No processes selected. Please go to Client Setup to select processes.")
        return

    if not risks:
        st.warning("No risks selected. Please go to Risk Selection to choose risks.")
        return

    # Matrix preview
    st.markdown("### Process-Risk Matrix")

    # Create matrix view
    matrix_data = []

    # Use assessments passed in (no re-fetch)
    assessment_lookup = {(a['process_id'], a['risk_id']): a for a in assessments}

    for proc in processes:
        row = {"Process": proc['custom_name'] or proc['process_name']}

        for risk in risks:
            # Check if assessment exists (using backend-aware lookup)
            assessment = assessment_lookup.get((proc['id'], risk['id']))

            if assessment:
                exposure = calculate_risk_exposure(
                    proc['criticality_per_day'],
                    assessment['vulnerability'],
                    assessment['resilience'],
                    assessment['expected_downtime'],
                    risk['probability']
                )
                row[risk['risk_name'][:20]] = f"✅ {format_currency(exposure, client['currency'])}"
            else:
                row[risk['risk_name'][:20]] = "⬜ Pending"

        matrix_data.append(row)

    df_matrix = pd.DataFrame(matrix_data)
    st.dataframe(df_matrix, width="stretch", hide_index=True)


def guided_assessment(client, processes, risks, assessments):
    """Guided step-by-step assessment interface."""
    st.subheader("🎯 Guided Assessment")

    if not st.session_state.current_client_id:
        return

    if not processes or not risks:
        st.warning("Please ensure you have selected both processes and risks.")
        return

    currency = client.get('currency', 'EUR')
    symbol = CURRENCY_SYMBOLS.get(currency, '€')

    # Build list of combinations (using data passed in, no re-fetch)
    assessment_lookup = {(a['process_id'], a['risk_id']): a for a in assessments}

    combinations = []
    for proc in processes:
        for risk in risks:
            existing = assessment_lookup.get((proc['id'], risk['id']))
            combinations.append({
                'process': proc,
                'risk': risk,
                'existing': existing,
                'completed': existing is not None
            })

    # Progress
    completed_count = sum(1 for c in combinations if c['completed'])
    st.progress(completed_count / len(combinations) if combinations else 0)
    st.write(f"Progress: {completed_count} / {len(combinations)} combinations assessed")

    # Select combination to assess
    pending = [c for c in combinations if not c['completed']]
    completed = [c for c in combinations if c['completed']]

    st.divider()

    # Tabs for pending vs editing
    tab1, tab2 = st.tabs(["📝 Pending", "✏️ Edit Completed"])

    with tab1:
        if not pending:
            st.success("🎉 All combinations have been assessed!")
        else:
            st.info(f"{len(pending)} combinations pending assessment")

            # Select next combination
            options = [f"{c['process']['process_name']} × {c['risk']['risk_name']}" for c in pending]
            selected_idx = st.selectbox(
                "Select combination to assess",
                options=range(len(options)),
                format_func=lambda x: options[x]
            )

            if selected_idx is not None:
                combo = pending[selected_idx]
                assessment_form(combo['process'], combo['risk'], None, client)

    with tab2:
        if not completed:
            st.info("No assessments completed yet")
        else:
            options = [f"{c['process']['process_name']} × {c['risk']['risk_name']}" for c in completed]
            selected_idx = st.selectbox(
                "Select combination to edit",
                options=range(len(options)),
                format_func=lambda x: options[x],
                key="edit_select"
            )

            if selected_idx is not None:
                combo = completed[selected_idx]
                assessment_form(combo['process'], combo['risk'], combo['existing'], client)


def assessment_form(process, risk, existing, client):
    """Assessment form for a single process-risk combination."""
    currency = client.get('currency', 'EUR')
    symbol = CURRENCY_SYMBOLS.get(currency, '€')

    st.markdown("---")

    # Process info
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### 📋 Process")
        st.write(f"**{process['process_name']}**")
        st.write(f"Criticality: {format_currency(process['criticality_per_day'], currency)}/day")

    with col2:
        st.markdown("### ⚠️ Risk")
        domain_color = get_domain_color(risk['domain'])
        st.markdown(f"**{risk['risk_name']}**")
        st.markdown(
            f"<span style='background-color:{domain_color}20; padding:2px 8px; border-radius:4px;'>"
            f"{get_domain_icon(risk['domain'])} {risk['domain']}</span>",
            unsafe_allow_html=True
        )
        st.write(f"Probability: {format_percentage(risk['probability'])}")

    st.divider()

    # Assessment inputs
    with st.form(f"assessment_{process['id']}_{risk['id']}"):
        st.markdown("### 📝 Assessment Inputs")

        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown("**Vulnerability**")
            st.caption("How likely is this process affected by this risk?")
            vulnerability = st.slider(
                "Vulnerability (%)",
                min_value=0,
                max_value=100,
                value=int(existing['vulnerability'] * 100) if existing else 0,
                step=5,
                key=f"vuln_{process['id']}_{risk['id']}",
                label_visibility="collapsed"
            )
            st.write(f"**{vulnerability}%**")

        with col2:
            st.markdown("**Resilience**")
            st.caption("How quickly can you recover from this risk?")
            resilience = st.slider(
                "Resilience (%)",
                min_value=0,
                max_value=100,
                value=int(existing['resilience'] * 100) if existing else 0,
                step=5,
                key=f"res_{process['id']}_{risk['id']}",
                label_visibility="collapsed"
            )
            st.write(f"**{resilience}%**")

        with col3:
            st.markdown("**Expected Downtime**")
            st.caption("Days until normal operations resume")
            downtime = st.number_input(
                "Downtime (days)",
                min_value=0,
                max_value=365,
                value=existing['expected_downtime'] if existing else 0,
                step=1,
                key=f"down_{process['id']}_{risk['id']}",
                label_visibility="collapsed"
            )
            st.write(f"**{downtime} days**")

        notes = st.text_area(
            "Notes (optional)",
            value=existing['notes'] if existing else "",
            placeholder="Add any notes about this assessment...",
            key=f"notes_{process['id']}_{risk['id']}"
        )

        # Preview calculation
        st.markdown("### 💰 Calculated Exposure")
        exposure = calculate_risk_exposure(
            process['criticality_per_day'],
            vulnerability / 100,
            resilience / 100,
            downtime,
            risk['probability']
        )

        st.metric(
            "Annual Risk Exposure",
            format_currency(exposure, currency),
            help="Criticality × Vulnerability × (1-Resilience) × Downtime × Probability"
        )

        # Formula breakdown
        with st.expander("See calculation breakdown"):
            st.markdown(f"""
            **PRISM Formula:**
            ```
            Exposure = Criticality × Vulnerability × (1-Resilience) × Downtime × Probability
            ```

            **Values:**
            - Criticality: {format_currency(process['criticality_per_day'], currency)}/day
            - Vulnerability: {vulnerability}%
            - Resilience: {resilience}% → Impact factor: {100-resilience}%
            - Downtime: {downtime} days
            - Probability: {format_percentage(risk['probability'])}

            **Calculation:**
            ```
            {symbol}{process['criticality_per_day']:,.0f} × {vulnerability/100:.2f} × {(100-resilience)/100:.2f} × {downtime} × {risk['probability']:.2f} = {symbol}{exposure:,.0f}/year
            ```
            """)

        submitted = st.form_submit_button("💾 Save Assessment", type="primary", width="stretch")

        if submitted:
            save_assessment(
                client_id=st.session_state.current_client_id,
                process_id=process['id'],
                risk_id=risk['id'],
                vulnerability=vulnerability / 100,
                resilience=resilience / 100,
                expected_downtime=downtime,
                notes=notes
            )
            st.success("✅ Assessment saved!")
            st.rerun()


def batch_assessment(client, processes, risks, assessments):
    """Batch assessment table interface."""
    st.subheader("📊 Batch Assessment")

    if not st.session_state.current_client_id:
        return

    if not processes or not risks:
        st.warning("Please ensure you have selected both processes and risks.")
        return

    st.markdown("""
    Enter assessments in bulk using the table below. Edit the values directly in the table, then click Save.
    """)

    # Use assessments passed in (no re-fetch)
    assessment_lookup = {(a['process_id'], a['risk_id']): a for a in assessments}

    data = []
    id_mapping = []  # Store process/risk IDs separately by row index

    for proc in processes:
        for risk in risks:
            existing = assessment_lookup.get((proc['id'], risk['id']))
            data.append({
                'Process': proc['process_name'][:30],
                'Risk': risk['risk_name'][:30],
                'Vulnerability (%)': int(existing['vulnerability'] * 100) if existing else 0,
                'Resilience (%)': int(existing['resilience'] * 100) if existing else 0,
                'Downtime (days)': existing['expected_downtime'] if existing else 0
            })

            # Keep track of IDs separately
            id_mapping.append({
                'proc_id': proc['id'],
                'risk_id': risk['id']
            })

    df = pd.DataFrame(data)

    # Editable dataframe - no hidden columns needed
    edited_df = st.data_editor(
        df,
        column_config={
            'Process': st.column_config.TextColumn(disabled=True),
            'Risk': st.column_config.TextColumn(disabled=True),
            'Vulnerability (%)': st.column_config.NumberColumn(min_value=0, max_value=100, step=5),
            'Resilience (%)': st.column_config.NumberColumn(min_value=0, max_value=100, step=5),
            'Downtime (days)': st.column_config.NumberColumn(min_value=0, max_value=365, step=1)
        },
        hide_index=True,
        width="stretch",
        num_rows="fixed"
    )

    if st.button("💾 Save All Assessments", type="primary", width="stretch"):
        saved_count = 0

        for idx, row in edited_df.iterrows():
            # Get IDs from our separate mapping using the row index
            ids = id_mapping[idx]

            save_assessment(
                client_id=st.session_state.current_client_id,
                process_id=ids['proc_id'],
                risk_id=ids['risk_id'],
                vulnerability=row['Vulnerability (%)'] / 100,
                resilience=row['Resilience (%)'] / 100,
                expected_downtime=int(row['Downtime (days)']),
                notes=""
            )
            saved_count += 1

        st.success(f"✅ Saved {saved_count} assessments!")
        st.rerun()  # Refresh to show updated values


def import_export_assessments(client, processes, risks, assessments):
    """Import/Export assessments via XLSX."""
    st.subheader("📥 Import / Export Assessments")

    if not st.session_state.current_client_id:
        st.warning("Please select a client to continue")
        return

    if not processes or not risks:
        st.warning("Please ensure you have selected both processes and risks.")
        return

    currency = client.get('currency', 'EUR')
    client_name_safe = client['name'].replace(' ', '_')

    col_dl, col_ul = st.columns(2)

    # --- Download template ---
    with col_dl:
        st.markdown("#### Download Assessment Template")
        st.markdown(
            "Download a spreadsheet with **all process-risk combinations**. "
            "Fill in the three assessment columns "
            "(**Vulnerability %**, **Resilience %**, **Downtime days**) "
            "and upload it back to import your values."
        )

        # Build export data
        assessment_lookup = {(a['process_id'], a['risk_id']): a for a in assessments}
        export_data = []

        for proc in processes:
            for risk in risks:
                existing = assessment_lookup.get((proc['id'], risk['id']))
                export_data.append({
                    'Process ID': proc['id'],
                    'Process Name': proc['process_name'],
                    'Risk ID': risk['id'],
                    'Risk Name': risk['risk_name'],
                    'Domain': risk['domain'],
                    'Criticality/Day': proc['criticality_per_day'],
                    'Probability (%)': round(risk['probability'] * 100, 2),
                    'Vulnerability (%)': int(existing['vulnerability'] * 100) if existing else '',
                    'Resilience (%)': int(existing['resilience'] * 100) if existing else '',
                    'Downtime (days)': existing['expected_downtime'] if existing else '',
                })

        df_export = pd.DataFrame(export_data)

        # Create formatted Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name='Assessments', index=False)
            ws = writer.sheets['Assessments']

            # --- Style definitions ---
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9'),
            )
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill_info = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_fill_input = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
            input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            zebra_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

            # Input columns (the ones the user fills): Vulnerability, Resilience, Downtime
            input_col_indices = []
            for col_idx, col_name in enumerate(df_export.columns, 1):
                if col_name in ('Vulnerability (%)', 'Resilience (%)', 'Downtime (days)'):
                    input_col_indices.append(col_idx)

            num_cols = len(df_export.columns)
            num_rows = len(df_export)

            # --- Format header row ---
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                if col_idx in input_col_indices:
                    cell.fill = header_fill_input
                else:
                    cell.fill = header_fill_info

            # --- Format data rows ---
            for row_idx in range(2, num_rows + 2):
                is_odd = (row_idx % 2) == 1
                for col_idx in range(1, num_cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')

                    if col_idx in input_col_indices:
                        # Input columns: yellow highlight so user knows what to fill
                        if cell.value == '' or cell.value is None:
                            cell.fill = input_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif is_odd:
                        cell.fill = zebra_fill

            # --- Auto-fit column widths ---
            for col_idx, col_name in enumerate(df_export.columns, 1):
                max_len = max(
                    len(str(col_name)),
                    df_export.iloc[:, col_idx - 1].astype(str).str.len().max()
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

            # --- Freeze header row ---
            ws.freeze_panes = 'A2'

            # --- Set row height for header ---
            ws.row_dimensions[1].height = 30

        output.seek(0)

        total_combos = len(export_data)
        filled = sum(1 for d in export_data if d['Vulnerability (%)'] != '')

        st.download_button(
            label=f"Download Template (.xlsx) -- {total_combos} combinations, {filled} filled",
            data=output.getvalue(),
            file_name=f"PRISM_Risk_Assessment_{client_name_safe}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # --- Upload completed template ---
    with col_ul:
        st.markdown("#### Upload Assessment File")
        st.markdown(
            "Upload the completed template. Rows with values in "
            "**Vulnerability**, **Resilience**, and **Downtime** columns "
            "will be imported and saved."
        )

        uploaded_file = st.file_uploader(
            "Upload completed assessment template",
            type=["xlsx"],
            key="assessment_upload",
        )

    # --- Process uploaded file (below both columns) ---
    if uploaded_file:
        try:
            df_import = pd.read_excel(uploaded_file)

            required_cols = ['Process ID', 'Risk ID', 'Vulnerability (%)', 'Resilience (%)', 'Downtime (days)']
            missing_cols = [col for col in required_cols if col not in df_import.columns]

            if missing_cols:
                st.error(f"Missing required columns: {', '.join(missing_cols)}. Please use the PRISM Assessment Template.")
            else:
                # Count rows that have assessment data filled in
                filled_rows = df_import[
                    df_import['Vulnerability (%)'].notna()
                    & df_import['Resilience (%)'].notna()
                    & df_import['Downtime (days)'].notna()
                ]

                st.info(f"Found **{len(filled_rows)}** rows with assessment data out of {len(df_import)} total rows.")

                if len(filled_rows) > 0:
                    # Show preview with only the key columns
                    preview_cols = ['Process ID', 'Risk ID', 'Vulnerability (%)', 'Resilience (%)', 'Downtime (days)']
                    if 'Process Name' in df_import.columns:
                        preview_cols = ['Process Name'] + preview_cols
                    st.dataframe(
                        filled_rows[preview_cols].reset_index(drop=True),
                        hide_index=True,
                        use_container_width=True,
                    )

                    if st.button("Apply Import", type="primary", key="apply_assessment_upload"):
                        saved_count = 0
                        error_count = 0

                        for idx, row in filled_rows.iterrows():
                            try:
                                save_assessment(
                                    client_id=st.session_state.current_client_id,
                                    process_id=row['Process ID'],
                                    risk_id=row['Risk ID'],
                                    vulnerability=float(row['Vulnerability (%)']) / 100,
                                    resilience=float(row['Resilience (%)']) / 100,
                                    expected_downtime=int(row['Downtime (days)']),
                                    notes=""
                                )
                                saved_count += 1
                            except Exception as e:
                                error_count += 1
                                st.warning(f"Row {idx + 2}: {str(e)}")

                        st.success(f"Imported {saved_count} assessments!")
                        if error_count > 0:
                            st.warning(f"{error_count} rows had errors")
                        st.rerun()
                else:
                    st.warning("No rows have all three assessment columns filled in (Vulnerability, Resilience, Downtime).")

        except Exception as e:
            st.error(f"Error reading file: {str(e)}")


def main():
    """Main page function."""
    st.title("🎯 Risk Assessment")
    st.markdown("Assess vulnerability, resilience, and downtime for each process-risk combination.")

    # Load ALL data once, pass to every function
    client, processes, risks, assessments = _load_page_data()

    client_selector_sidebar(processes, risks, assessments)

    if not st.session_state.current_client_id:
        st.warning("Please select a client to continue")
        if st.button("Go to Client Setup"):
            st.switch_page("pages/1_Client_Setup.py")
        return

    # Tabs
    tab1, tab2, tab3 = st.tabs([
        "📋 Batch Mode",
        "🎯 Guided Assessment",
        "📥 Import / Export"
    ])

    with tab1:
        batch_assessment(client, processes, risks, assessments)

    with tab2:
        guided_assessment(client, processes, risks, assessments)

    with tab3:
        import_export_assessments(client, processes, risks, assessments)

    # Navigation
    st.divider()
    col1, col2, col3 = st.columns([1, 2, 1])

    with col1:
        if st.button("← Risk Selection"):
            st.switch_page("pages/3_Risk_Selection.py")

    with col3:
        if assessments:
            if st.button("Next: Results →", type="primary"):
                st.switch_page("pages/5_Results_Dashboard.py")


if __name__ == "__main__":
    main()
