"""PRISM Brain — Data Sources & Engine Status
==============================================
Monitor the probability engine, check data source credentials,
trigger recomputation, and view historical computation runs.
"""

import io
import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
from utils.theme import inject_prism_theme

import sys
APP_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(APP_DIR))

from modules.database import is_backend_online
from modules.api_client import (
    api_engine_status,
    api_engine_compute_all,
    api_engine_history_runs,
    api_engine_history_run_detail,
    api_engine_history_compare,
    clear_cache,
)

st.set_page_config(
    page_title="Data Sources | PRISM Brain",
    page_icon="📡",
    layout="wide",
)

inject_prism_theme()


def show_engine_status():
    """Display engine health, version, and credential status."""
    st.subheader("Engine Status")

    backend_online = is_backend_online()
    if backend_online:
        st.success("Backend connected and operational")
    else:
        st.error("Backend offline — start the server with `python main.py`")
        return

    status = api_engine_status(use_cache=False)
    if not status:
        st.warning("Could not reach the engine status endpoint.")
        return

    col1, col2, col3, col4 = st.columns(4, gap="medium")
    with col1:
        st.metric("Engine Version", status.get("engine_version", "—"))
    with col2:
        st.metric("Spec Version", status.get("spec_version", "—"))
    with col3:
        st.metric("Total Events", status.get("total_events", "—"))
    with col4:
        st.metric("Fallback Rates Loaded", status.get("fallback_rates_loaded", "—"))

    credentials = status.get("api_credentials", {})
    no_key_sources = status.get("no_key_sources", [])

    if credentials:
        with st.expander("API Credentials", expanded=False):
            cred_rows = []
            for source, available in credentials.items():
                cred_rows.append({
                    "Source": source,
                    "Status": "Configured" if available else "Missing",
                    "Icon": "✅" if available else "❌",
                })
            df = pd.DataFrame(cred_rows)
            st.dataframe(df, width=600, hide_index=True)

            if no_key_sources:
                st.caption(f"Sources that require no API key: {', '.join(no_key_sources)}")


def show_compute_section():
    """Buttons to trigger engine computation for all 174 events."""
    st.subheader("Compute Probabilities")

    st.info(
        "Click below to compute dynamic probabilities for all 174 risk events "
        "using the PRISM probability engine. The engine fetches live data from "
        "external APIs (EM-DAT, FRED, World Bank, etc.) and applies the three-method "
        "framework (A / B / C) to produce calibrated probabilities. "
        "Results are automatically archived for historical tracking."
    )

    col1, col2 = st.columns([1, 1], gap="medium")

    with col1:
        compute_btn = st.button(
            "Compute All 174 Events",
            key="compute_all_btn",
            type="primary",
            use_container_width=True,
        )

    with col2:
        if st.button("Clear Cache", key="clear_cache_btn", use_container_width=True):
            clear_cache()
            st.success("Cache cleared. Next computation will fetch fresh data.")

    if compute_btn:
        progress = st.empty()
        result_area = st.empty()

        with progress.container():
            st.info("Computing probabilities for all 174 events — this may take up to 2 minutes on first run...")

        result = api_engine_compute_all(use_cache=False)

        progress.empty()

        if result is None:
            result_area.error("Engine computation failed. Check that the backend is running.")
            return

        with result_area.container():
            st.success(f"Computed {len(result)} events successfully!")

            methods = {"A": 0, "B": 0, "C": 0, "FALLBACK": 0}
            probs = []
            for eid, r in result.items():
                if isinstance(r, dict):
                    m = r.get("layer1", {}).get("method", "FALLBACK")
                    methods[m] = methods.get(m, 0) + 1
                    p = r.get("layer1", {}).get("p_global")
                    if p is not None:
                        probs.append(p)

            met_cols = st.columns(4, gap="medium")
            labels = {
                "A": "Method A (Frequency)",
                "B": "Method B (Survey)",
                "C": "Method C (Structural)",
                "FALLBACK": "Fallback Only",
            }
            for i, (key, label) in enumerate(labels.items()):
                with met_cols[i]:
                    st.metric(label, methods.get(key, 0))

            if probs:
                bins = {"0-25%": 0, "25-50%": 0, "50-75%": 0, "75-100%": 0}
                for p in probs:
                    pct = p * 100 if p <= 1 else p
                    if pct < 25:
                        bins["0-25%"] += 1
                    elif pct < 50:
                        bins["25-50%"] += 1
                    elif pct < 75:
                        bins["50-75%"] += 1
                    else:
                        bins["75-100%"] += 1
                st.write("**Probability Distribution:**")
                st.bar_chart(pd.Series(bins, name="Events"))

            st.caption(f"Computed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


def _generate_history_excel(calculation_id: str, snapshots: list, run_info: dict) -> bytes:
    """Generate a styled Excel workbook for a historical compute run."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    PRISM_BLUE = "1F4E79"
    PRISM_LIGHT = "D6E4F0"
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color=PRISM_BLUE, end_color=PRISM_BLUE, fill_type="solid")
    ALT_FILL = PatternFill(start_color=PRISM_LIGHT, end_color=PRISM_LIGHT, fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    wb = Workbook()

    # --- Tab 1: Run Summary ---
    ws1 = wb.active
    ws1.title = "Run Summary"
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 40

    summary_rows = [
        ("Calculation ID", calculation_id),
        ("Date", run_info.get("start_time", "")[:19] if run_info.get("start_time") else ""),
        ("Events Processed", run_info.get("events_processed", len(snapshots))),
        ("Events Succeeded", run_info.get("events_succeeded", len(snapshots))),
        ("Duration (seconds)", run_info.get("duration_seconds", "")),
        ("Status", run_info.get("status", "")),
        ("Trigger", run_info.get("trigger", "")),
    ]

    # Method distribution
    method_counts = {}
    for s in snapshots:
        m = s.get("method", "?")
        method_counts[m] = method_counts.get(m, 0) + 1

    for row_idx, (label, value) in enumerate(summary_rows, 1):
        cell_a = ws1.cell(row=row_idx, column=1, value=label)
        cell_b = ws1.cell(row=row_idx, column=2, value=value)
        cell_a.font = Font(bold=True)

    row = len(summary_rows) + 2
    ws1.cell(row=row, column=1, value="Method Distribution").font = Font(bold=True)
    for method, count in sorted(method_counts.items()):
        row += 1
        ws1.cell(row=row, column=1, value=f"Method {method}")
        ws1.cell(row=row, column=2, value=count)

    # --- Tab 2: All Event Probabilities ---
    ws2 = wb.create_sheet("Event Probabilities")
    headers = [
        "Event ID", "Event Name", "Domain", "Family", "P_global (%)",
        "Prior", "Method", "Data Source", "Dynamic", "Modifiers", "Confidence"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    conf_map = {0.9: "High", 0.6: "Medium", 0.3: "Low"}

    for row_idx, s in enumerate(snapshots, 2):
        values = [
            s.get("event_id", ""),
            s.get("event_name", ""),
            s.get("domain", ""),
            s.get("family", ""),
            s.get("probability_pct"),
            round(s["prior"], 4) if s.get("prior") is not None else None,
            s.get("method", ""),
            s.get("data_source", ""),
            "Yes" if s.get("is_dynamic") else "No",
            s.get("modifier_count", 0),
            conf_map.get(s.get("confidence_score"), "Medium"),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # Auto-width
    for col in ws2.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws2.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # --- Tab 3: By Domain ---
    ws3 = wb.create_sheet("By Domain")
    domain_headers = ["Domain", "Events", "Method A", "Method B", "Method C", "Avg P_global (%)"]
    for col_idx, header in enumerate(domain_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    domain_stats = {}
    for s in snapshots:
        d = s.get("domain", "Unknown")
        if d not in domain_stats:
            domain_stats[d] = {"count": 0, "methods": {"A": 0, "B": 0, "C": 0}, "probs": []}
        domain_stats[d]["count"] += 1
        m = s.get("method", "")
        if m in domain_stats[d]["methods"]:
            domain_stats[d]["methods"][m] += 1
        if s.get("probability_pct") is not None:
            domain_stats[d]["probs"].append(s["probability_pct"])

    for row_idx, (domain, stats) in enumerate(sorted(domain_stats.items()), 2):
        avg_p = sum(stats["probs"]) / len(stats["probs"]) if stats["probs"] else 0
        values = [
            domain,
            stats["count"],
            stats["methods"]["A"],
            stats["methods"]["B"],
            stats["methods"]["C"],
            round(avg_p, 2),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    for col in ws3.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws3.column_dimensions[col_letter].width = min(max_len + 3, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def show_history_tab():
    """Display probability computation history with drill-down and export."""
    st.subheader("Computation History")

    run_data = api_engine_history_runs(limit=50)
    if not run_data or not run_data.get("runs"):
        st.info(
            "No computation history yet. "
            "Go to the **Compute Probabilities** tab and run a computation to start building history."
        )
        return

    runs = run_data["runs"]

    # --- Run list table ---
    run_df = pd.DataFrame(runs)
    run_df["start_time"] = pd.to_datetime(run_df["start_time"])
    run_df["Date"] = run_df["start_time"].dt.strftime("%Y-%m-%d %H:%M")
    run_df["Duration"] = run_df["duration_seconds"].apply(lambda x: f"{x:.1f}s" if x else "—")
    run_df["Events"] = run_df["events_succeeded"].astype(str) + "/" + run_df["events_processed"].astype(str)

    display_df = run_df[["Date", "calculation_id", "Events", "Duration", "status", "trigger"]].copy()
    display_df.columns = ["Date", "Run ID", "Events", "Duration", "Status", "Trigger"]

    st.dataframe(display_df, use_container_width=True, hide_index=True)

    # --- Drill into a specific run ---
    st.divider()
    st.subheader("Run Details")

    col1, col2 = st.columns(2)

    run_options = [r["calculation_id"] for r in runs]
    run_labels = {
        r["calculation_id"]: f"{r['calculation_id']} — {r['start_time'][:16] if r.get('start_time') else '?'}"
        for r in runs
    }

    with col1:
        selected_run = st.selectbox(
            "Select a run to view details",
            options=run_options,
            format_func=lambda x: run_labels.get(x, x),
        )

    if selected_run:
        detail = api_engine_history_run_detail(selected_run)
        if detail and detail.get("snapshots"):
            snapshots = detail["snapshots"]

            st.markdown(f"**Run {selected_run}** — {len(snapshots)} events")

            snap_df = pd.DataFrame(snapshots)

            # Domain filter
            with col2:
                domains = sorted(snap_df["domain"].dropna().unique())
                domain_filter = st.selectbox("Filter by Domain", ["All"] + list(domains))

            if domain_filter != "All":
                snap_df = snap_df[snap_df["domain"] == domain_filter]

            # Confidence label
            conf_map = {0.9: "High", 0.6: "Medium", 0.3: "Low"}
            snap_df["Confidence"] = snap_df["confidence_score"].map(
                lambda x: conf_map.get(x, "Medium") if x else "Medium"
            )
            snap_df["Dynamic"] = snap_df["is_dynamic"].map(lambda x: "Yes" if x else "No")

            display_snap = snap_df[[
                "event_id", "event_name", "domain", "probability_pct",
                "prior", "method", "data_source", "Dynamic", "modifier_count", "Confidence"
            ]].copy()
            display_snap.columns = [
                "Event ID", "Event Name", "Domain", "P_global (%)",
                "Prior", "Method", "Data Source", "Dynamic", "Modifiers", "Confidence"
            ]

            st.dataframe(
                display_snap,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "P_global (%)": st.column_config.NumberColumn(format="%.2f"),
                    "Prior": st.column_config.NumberColumn(format="%.4f"),
                },
            )

            # Find the run info for Excel export
            run_info = next((r for r in runs if r["calculation_id"] == selected_run), {})

            excel_data = _generate_history_excel(selected_run, snapshots, run_info)
            st.download_button(
                label="Download Run as Excel",
                data=excel_data,
                file_name=f"PRISM_History_{selected_run}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.warning("No snapshot data found for this run.")

    # --- Compare two runs ---
    if len(runs) >= 2:
        st.divider()
        st.subheader("Compare Runs")

        comp_col1, comp_col2 = st.columns(2)
        with comp_col1:
            run_a = st.selectbox(
                "Run A (older)",
                options=run_options,
                format_func=lambda x: run_labels.get(x, x),
                index=min(1, len(run_options) - 1),
                key="comp_a",
            )
        with comp_col2:
            run_b = st.selectbox(
                "Run B (newer)",
                options=run_options,
                format_func=lambda x: run_labels.get(x, x),
                index=0,
                key="comp_b",
            )

        if st.button("Compare Runs", type="primary") and run_a != run_b:
            comp = api_engine_history_compare(run_a, run_b)
            if comp and comp.get("comparison"):
                comp_df = pd.DataFrame(comp["comparison"])
                comp_df = comp_df.sort_values("delta", key=abs, ascending=False, na_position="last")

                # Summary metrics
                changes = comp_df["delta"].dropna()
                if len(changes) > 0:
                    mc1, mc2, mc3 = st.columns(3)
                    with mc1:
                        st.metric("Events Changed", f"{(changes.abs() > 0.01).sum()} / {len(comp_df)}")
                    with mc2:
                        st.metric("Avg Change", f"{changes.mean():+.2f} pp")
                    with mc3:
                        st.metric("Max Change", f"{changes.abs().max():.2f} pp")

                st.dataframe(
                    comp_df[["event_id", "event_name", "domain", "prob_a", "prob_b", "delta", "method_a", "method_b"]],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "event_id": "Event ID",
                        "event_name": "Event Name",
                        "domain": "Domain",
                        "prob_a": st.column_config.NumberColumn("Run A (%)", format="%.2f"),
                        "prob_b": st.column_config.NumberColumn("Run B (%)", format="%.2f"),
                        "delta": st.column_config.NumberColumn("Change (pp)", format="%+.2f"),
                        "method_a": "Method A",
                        "method_b": "Method B",
                    },
                )
            else:
                st.warning("Could not compare runs. Check that both runs have data.")
        elif run_a == run_b:
            st.caption("Select two different runs to compare.")


def main():
    st.title("Data Sources & Engine")
    st.markdown(
        "Monitor engine health, compute probabilities, and view historical computation runs."
    )

    tab1, tab2, tab3 = st.tabs([
        "Engine Status",
        "Compute Probabilities",
        "History Archive",
    ])

    with tab1:
        show_engine_status()

    with tab2:
        show_compute_section()

    with tab3:
        show_history_tab()

    st.divider()

    # Navigation
    nav_col1, _, nav_col3 = st.columns([1, 2, 1], gap="large")
    with nav_col1:
        if st.button("← Results Dashboard", use_container_width=True, key="nav_results"):
            st.switch_page("pages/5_Results_Dashboard.py")
    with nav_col3:
        if st.button("Risk Selection →", use_container_width=True, key="nav_risk"):
            st.switch_page("pages/3_Risk_Selection.py")


if __name__ == "__main__":
    main()
